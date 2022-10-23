# -*- coding: utf-8 -*-
"""
Created on Sun Oct 23 17:20:19 2022

@author: user
"""

import openpyxl
import statistics as stats
from datetime import datetime

# load IBM data
wb = openpyxl.load_workbook('Stocks.xlsx')
IBMsheet = wb.get_sheet_by_name('IBM')

# all 2010 data
#start = datetime.datetime(2010, 1, 1)
#end = datetime.datetime(2010, 12, 31)

#start_date = start.strftime("%m%d%Y")
#end_date = end.strftime("%m%d%Y")

#for row in IBMsheet.iter_rows(min_row=1, min_col=1):
 #   for cell in row:
  #      if start_date <= cell.value <= end_date:
   #         print(IBMsheet.cell(row=cell.row, column=1).value)
    #        print(IBMsheet.cell(row=cell.row, column=2).value)
     #       print(IBMsheet.cell(row=cell.row, column=3).value)
      #      print(IBMsheet.cell(row=cell.row, column=4).value)
       #     print(IBMsheet.cell(row=cell.row, column=5).value)
        #    print(IBMsheet.cell(row=cell.row, column=6).value)
         #   print(IBMsheet.cell(row=cell.row, column=7).value)
            
# min max and mean of 2010 prices (col D)
values = []

for row in IBMsheet.iter_rows(min_row=1, min_col=1):
    for cell in row:
        if cell.value.year == 2010:
            values.append(IBMsheet.cell(row=cell.row, column=4).value)
            
            
print("Minimum price: {}".format(min(values)))
print("Maximum price: {}".format(max(values)))
print("Mean priceL {}".format(stats.mean(values)))


